import os
import cv2
import time
import threading
import smtplib
import subprocess
from email.message import EmailMessage
from ultralytics import YOLO
from twilio.rest import Client
from dotenv import load_dotenv
from datetime import datetime
from openpyxl import Workbook, load_workbook

# =====================================
# LOAD ENV VARIABLES
# =====================================
load_dotenv("email.env")

sender_email = os.getenv("SENDER_EMAIL")
app_password = os.getenv("APP_PASSWORD")
receiver_email = os.getenv("RECEIVER_EMAIL")

account_sid = os.getenv("TWILIO_ACCOUNT_SID")
auth_token = os.getenv("TWILIO_AUTH_TOKEN")

twilio_from = os.getenv("TWILIO_FROM")
twilio_to = os.getenv("TWILIO_TO")

client = Client(account_sid, auth_token)

# =====================================
# SETTINGS
# =====================================
VIDEO_DURATION = 10
COOLDOWN_TIME = 15


last_alert_time = 0
frame_count = 0

# =====================================
# MASK FUNCTIONS
# =====================================
def mask_email(email):
    if not email:
        return "Not Configured"

    if "@" not in email:
        return "Invalid Email"

    name, domain = email.split("@")

    if len(name) <= 2:
        masked = name[0] + "x" * (len(name) - 1)
    else:
        masked = name[0] + "x" * (len(name) - 2) + name[-1]

    return masked + "@" + domain


def mask_phone(number):
    if not number:
        return "Not Configured"

    number = number.replace("whatsapp:+", "")

    if len(number) <= 2:
        return number[0] + "x" * (len(number) - 1)

    return number[0] + "x" * (len(number) - 2) + number[-1]


# =====================================
# LOAD YOLO MODEL
# =====================================
print("Loading YOLO model...")

model = YOLO("yolov8n.pt")

try:
    model.to("cuda")
    print("GPU Enabled")
except:
    print("Running on CPU")


# =====================================
# CAMERA
# =====================================
cap = cv2.VideoCapture(0)

if not cap.isOpened():
    print("ERROR: Camera not found")
    exit()

cap.set(cv2.CAP_PROP_FRAME_WIDTH, 640)
cap.set(cv2.CAP_PROP_FRAME_HEIGHT, 480)

# =====================================
# NIGHT ENHANCEMENT
# =====================================
def night_enhancement(frame):

    gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

    if gray.mean() < 60:
        frame = cv2.convertScaleAbs(
            frame,
            alpha=1.5,
            beta=40
        )

    return frame
# =====================================
# EMAIL ALERT
# =====================================
def send_email(image_path, video_path):
    try:
        msg = EmailMessage()

        msg["Subject"] = "🚨 AI CCTV ALERT"
        msg["From"] = sender_email
        msg["To"] = receiver_email

        msg.set_content(
            "Person detected by AI CCTV system."
        )

        with open(image_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="image",
                subtype="jpeg",
                filename="alert.jpg"
            )

        with open(video_path, "rb") as f:
            msg.add_attachment(
                f.read(),
                maintype="video",
                subtype="mp4",
                filename="alert.mp4"
            )

        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as smtp:
            smtp.login(sender_email, app_password)
            smtp.send_message(msg)

        print(f"Email sent -> {mask_email(receiver_email)}")

    except Exception as e:
        print("Email Error:", e)

# =====================================
# WHATSAPP ALERT
# =====================================
def send_whatsapp():
    try:
        message = client.messages.create(
            body="🚨 ALERT! Person detected by AI CCTV",
            from_=twilio_from,
            to=twilio_to
        )

        print("SUCCESS")
        print("SID:", message.sid)

    except Exception as e:
        print("WhatsApp Error:", e)


# =====================================
# RECORD VIDEO
# =====================================
def record_video(cap, output_path):

    fourcc = cv2.VideoWriter_fourcc(*"mp4v")

    out = cv2.VideoWriter(
        output_path,
        fourcc,
        20.0,
        (640, 480)
    )

    start_time = time.time()

    while (
        time.time() - start_time
        < VIDEO_DURATION
    ):
        ret, frame = cap.read()

        if not ret:
            break

        timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        cv2.putText(
            frame,
            timestamp,
            (10, 30),
            cv2.FONT_HERSHEY_SIMPLEX,
            0.7,
            (0, 255, 255),
            2
        )

        out.write(frame)

    out.release()

def record_video_ffmpeg(output_path):

    cmd = [
        r"C:\Users\HP\Downloads\ffmpeg-8.1.1-essentials_build\ffmpeg-8.1.1-essentials_build\bin\ffmpeg.exe",
    
        "-f", "dshow",
        "-rtbufsize", "512M",
        "-i", "video=HP HD Camera:audio=Microphone (Realtek(R) Audio)",
        "-t", str(VIDEO_DURATION),
    
        "-vf",
        "drawtext=text='%{localtime}':x=10:y=10:fontsize=24:fontcolor=yellow",
    
        "-c:v", "libx264",
        "-pix_fmt", "yuv420p",
        "-c:a", "aac",
    
        "-y",
        output_path
    ]

    subprocess.run(cmd)

def save_detection_to_excel(person_count, image_name, video_name):

    excel_file = "detection_history.xlsx"

    if not os.path.exists(excel_file):

        wb = Workbook()
        ws = wb.active

        ws.append([
            "Rank",
            "Detection Time",
            "Person Count",
            "Alert Sent",
            "Image",
            "Video"
        ])

        wb.save(excel_file)

    wb = load_workbook(excel_file)
    ws = wb.active

    rank = ws.max_row

    ws.append([
        rank,
        datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        person_count,
        "Yes",
        image_name,
        video_name
    ])

    wb.save(excel_file)
# =====================================
# ALERT SYSTEM
# =====================================
def alert_system(frame, person_count):

    global cap

    image_path = "alert.jpg"
    final_video = "alert.mp4"
    
    save_detection_to_excel(
    person_count,
    image_path,
    final_video
    )
    
    cv2.imwrite(image_path, frame)

    threading.Thread(
        target=send_whatsapp,
        daemon=True
    ).start()
    
    print("Camera Released...")
    cap.release()
    
    time.sleep(2)
    
    print("Recording 10 sec video with audio...")
    record_video_ffmpeg(final_video)
    
    print("Camera Reopened...")
    cap = cv2.VideoCapture(0)
    
    threading.Thread(
        target=send_email,
        args=(image_path, final_video),
        daemon=True
    ).start()
   

# =====================================
# MAIN LOOP
# =====================================
print("AI CCTV Started...")
print("Press Q to Exit")
def get_status():
    try:
        with open("status.txt", "r") as f:
            return f.read().strip()
    except:
        return "ON"

while True:
    if get_status() == "OFF":
        time.sleep(1)
        continue
    

    ret, frame = cap.read()

    if not ret:
        time.sleep(1)
        continue
    
     
    frame = night_enhancement(frame)

    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    cv2.putText(
        frame,
        timestamp,
        (10, 30),
        cv2.FONT_HERSHEY_SIMPLEX,
        0.7,
        (0, 255, 255),
        2
    )

    frame_count += 1

    # Skip frames
    if frame_count % 3 != 0:
        continue

    person_detected = False
    person_count = 0

    # YOLO Detection
    results = model(
        frame,
        classes=[0],      # person only
        imgsz=320,
        conf=0.25,
        verbose=False
    )

    for result in results:

        if result.boxes is None:
            continue

        for box in result.boxes:
            
            person_count += 1
            person_detected = True

            x1, y1, x2, y2 = map(
                int,
                box.xyxy[0]
            )

            confidence = float(
                box.conf[0]
            )

            cv2.rectangle(
                frame,
                (x1, y1),
                (x2, y2),
                (0, 255, 0),
                2
            )

            cv2.putText(
                frame,
                f"PERSON {confidence:.2f}",
                (x1, y1 - 10),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.6,
                (0, 255, 0),
                2
            )
            cv2.putText(
                frame,
                f"Persons: {person_count}",
                (10, 60),
                cv2.FONT_HERSHEY_SIMPLEX,
                0.8,
                (0, 0, 255),
                2
             )


    cv2.imshow(
        "AI CCTV Surveillance",
        frame
    )

    current_time = time.time()

    if (
        person_detected
        and
        current_time - last_alert_time
        > COOLDOWN_TIME
    ):

        print(
            "Person Detected!"
        )

        print(
            "Sending Alerts..."
        )

        threading.Thread(
            target=alert_system,
            args=(frame.copy(),person_count)
        ).start()

        last_alert_time = current_time

    if (
        cv2.waitKey(1)
        & 0xFF
        == ord("q")
    ):
        break


# =====================================
# CLEANUP
# =====================================
cap.release()
cv2.destroyAllWindows()

print("System Stopped")
